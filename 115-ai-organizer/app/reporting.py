from __future__ import annotations

import html
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Settings
from .db import db_session, file_stats, init_db, list_plans
from .operations import AUTO_ORGANIZE_CATEGORIES


REPORT_COLUMNS = (
    "计划ID",
    "批准状态",
    "可自动执行",
    "风险提示",
    "分类",
    "置信度",
    "原文件名",
    "建议文件名",
    "原路径",
    "建议路径",
    "大小",
    "SHA1",
    "115文件ID",
    "判断依据",
    "执行状态",
)


def _format_size(value: int | None) -> str:
    size = float(value or 0)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024 or unit == "TB":
            return f"{int(size)} B" if unit == "B" else f"{size:.2f} {unit}"
        size /= 1024
    return "0 B"


def _duplicate_map(plans: list[dict[str, Any]]) -> dict[int, list[int]]:
    groups: dict[tuple[str, int], list[int]] = defaultdict(list)
    for row in plans:
        sha1 = str(row.get("hash_sha1") or "").strip().lower()
        size = int(row.get("size") or 0)
        if sha1 and size > 0:
            groups[(sha1, size)].append(int(row["id"]))
    result: dict[int, list[int]] = {}
    for ids in groups.values():
        if len(ids) > 1:
            for plan_id in ids:
                result[plan_id] = ids
    return result


def _report_row(
    row: dict[str, Any], duplicate_ids: dict[int, list[int]]
) -> dict[str, Any]:
    plan_id = int(row["id"])
    risks: list[str] = []
    if not row.get("file_id") or row.get("file_id_source") != "native":
        risks.append("缺少115原生ID")
    if row.get("confidence") == "low":
        risks.append("低置信度")
    if row.get("category") == "待识别":
        risks.append("待人工识别")
    if row.get("category") not in AUTO_ORGANIZE_CATEGORIES:
        risks.append("附件或其他类型需关联主文件")
    if plan_id in duplicate_ids:
        risks.append(f"疑似重复组({len(duplicate_ids[plan_id])}项)")
    executable = not risks and bool(row.get("approved"))
    return {
        "计划ID": plan_id,
        "批准状态": "已批准" if row.get("approved") else "未批准",
        "可自动执行": "是" if executable else "否",
        "风险提示": "；".join(risks),
        "分类": row.get("category") or "",
        "置信度": row.get("confidence") or "",
        "原文件名": row.get("original_name") or "",
        "建议文件名": row.get("suggested_name") or "",
        "原路径": row.get("original_path") or "",
        "建议路径": row.get("suggested_path") or "",
        "大小": _format_size(row.get("size")),
        "SHA1": row.get("hash_sha1") or "",
        "115文件ID": row.get("file_id") or "",
        "判断依据": row.get("reason") or "",
        "执行状态": row.get("execute_status") or "",
    }


def collect_report(settings: Settings) -> dict[str, Any]:
    init_db(settings.db_path)
    with db_session(settings.db_path) as conn:
        current = conn.execute(
            "SELECT started_at FROM scan_runs WHERE status = 'ok' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        snapshot_start = current["started_at"] if current else ""
        plans = [dict(row) for row in conn.execute(
            """
            SELECT p.*, f.size, f.hash_sha1, f.file_id_source, f.parent_id
            FROM organize_plans p
            JOIN files f ON f.id = p.file_row_id
            WHERE f.scan_time >= ?
            ORDER BY p.id
            """,
            (snapshot_start,),
        )]
        stats = file_stats(conn)
    duplicate_ids = _duplicate_map(plans)
    rows = [_report_row(row, duplicate_ids) for row in plans]
    duplicate_groups = {
        tuple(ids) for ids in duplicate_ids.values()
    }
    risk_counts = Counter()
    for row in rows:
        for risk in str(row["风险提示"]).split("；"):
            if risk:
                risk_counts[risk.split("(", 1)[0]] += 1
    return {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "stats": stats,
        "rows": rows,
        "duplicate_group_count": len(duplicate_groups),
        "duplicate_file_count": len(duplicate_ids),
        "approved_count": sum(1 for row in rows if row["批准状态"] == "已批准"),
        "executable_count": sum(1 for row in rows if row["可自动执行"] == "是"),
        "risk_counts": dict(risk_counts),
    }


def _write_excel(report: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"
    stats = report["stats"]
    overview_rows = [
        ("生成时间", report["generated_at"]),
        ("文件数", stats["file_count"]),
        ("文件夹数", stats["folder_count"]),
        ("总容量", _format_size(stats["total_size"])),
        ("待识别", stats["pending_count"]),
        ("疑似重复组", report["duplicate_group_count"]),
        ("疑似重复文件", report["duplicate_file_count"]),
        ("已批准", report["approved_count"]),
        ("当前可执行", report["executable_count"]),
        ("说明", "报告只提供建议；疑似重复文件不会自动删除。"),
    ]
    for item in overview_rows:
        overview.append(item)
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 72
    overview["A1"].font = Font(bold=True)

    detail = workbook.create_sheet("整理计划")
    detail.append(REPORT_COLUMNS)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in report["rows"]:
        detail.append([row[column] for column in REPORT_COLUMNS])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    widths = [10, 11, 12, 24, 12, 10, 38, 38, 60, 60, 14, 42, 24, 50, 16]
    for index, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(index)].width = width
    for row in detail.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[13].alignment = Alignment(wrap_text=True, vertical="top")
        if row[3].value:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")

    workbook.save(path)


def _write_html(report: dict[str, Any], path: Path) -> None:
    stats = report["stats"]
    category_text = " · ".join(
        f"{html.escape(str(name))} {count}"
        for name, count in sorted((stats.get("categories") or {}).items())
    ) or "尚无分类"
    cards = (
        ("文件", stats["file_count"]),
        ("容量", _format_size(stats["total_size"])),
        ("待识别", stats["pending_count"]),
        ("重复风险", report["duplicate_file_count"]),
        ("已批准", report["approved_count"]),
        ("可执行", report["executable_count"]),
    )
    card_html = "".join(
        f'<div class="card"><span>{html.escape(str(label))}</span><b>{html.escape(str(value))}</b></div>'
        for label, value in cards
    )
    row_html: list[str] = []
    for row in report["rows"]:
        risk = str(row["风险提示"])
        css = "risk" if risk else "safe"
        values = [row[column] for column in REPORT_COLUMNS]
        row_html.append(
            f'<tr class="{css}">' + "".join(
                f"<td>{html.escape(str(value or ''))}</td>" for value in values
            ) + "</tr>"
        )
    headers = "".join(f"<th>{html.escape(column)}</th>" for column in REPORT_COLUMNS)
    payload = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>115 文件整理报告</title>
<style>
body{{font-family:"Microsoft YaHei",sans-serif;margin:0;background:#f5f7fa;color:#1f2937}}
.wrap{{max-width:1600px;margin:auto;padding:20px}} h1{{margin:0 0 6px;font-size:24px}}
.muted{{color:#64748b;font-size:13px}} .cards{{display:grid;grid-template-columns:repeat(6,minmax(110px,1fr));gap:10px;margin:16px 0}}
.card{{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px}} .card span{{display:block;color:#64748b;font-size:12px}} .card b{{font-size:20px}}
.panel{{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px;margin-bottom:12px;overflow:auto}}
table{{border-collapse:collapse;width:100%;font-size:12px}} th,td{{border-bottom:1px solid #e5e7eb;padding:7px;text-align:left;vertical-align:top;max-width:360px}}
th{{position:sticky;top:0;background:#1f4e78;color:white;white-space:nowrap}} tr.risk{{background:#fffaf0}} tr.safe{{background:#f4fff7}}
@media(max-width:900px){{.cards{{grid-template-columns:repeat(2,1fr)}}}}
</style></head><body><div class="wrap">
<h1>115 文件整理报告</h1><div class="muted">生成于 {html.escape(report['generated_at'])}。绿色表示已批准且无已知风险；黄色必须人工复核。程序永不自动删除。</div>
<div class="cards">{card_html}</div>
<div class="panel"><b>分类：</b>{category_text}</div>
<div class="panel"><table><thead><tr>{headers}</tr></thead><tbody>{''.join(row_html)}</tbody></table></div>
</div></body></html>"""
    path.write_text(payload, encoding="utf-8")


def export_reports(settings: Settings, output_dir: str | Path) -> dict[str, Any]:
    report = collect_report(settings)
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = target / f"115_整理报告_{stamp}.html"
    xlsx_path = target / f"115_整理报告_{stamp}.xlsx"
    json_path = target / f"115_整理报告_{stamp}.json"
    _write_html(report, html_path)
    _write_excel(report, xlsx_path)
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "html": str(html_path.resolve()),
        "xlsx": str(xlsx_path.resolve()),
        "json": str(json_path.resolve()),
        "summary": {
            "file_count": report["stats"]["file_count"],
            "duplicate_group_count": report["duplicate_group_count"],
            "duplicate_file_count": report["duplicate_file_count"],
            "approved_count": report["approved_count"],
            "executable_count": report["executable_count"],
        },
    }
