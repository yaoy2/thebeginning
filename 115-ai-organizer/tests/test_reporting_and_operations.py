import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.config import load_settings
from app.db import connect, set_plan_approved
from app.operations import (
    OperationError,
    approve_safe_plans,
    build_manifest,
    confirmation_code,
    execute_manifest,
    load_manifest,
    save_manifest,
)
from app.reporting import export_reports
from app.scanner import scan


def fake_tree(path: str):
    catalog = {
        "/云下载": [
            {
                "name": "流浪地球2.2023.1080p.mkv",
                "file_id": "file-1",
                "parent_id": "123",
                "is_dir": False,
                "size": 100,
                "hash_info": {"sha1": "same-hash"},
            },
            {
                "name": "流浪地球2.2023.1080p.副本.mkv",
                "file_id": "file-2",
                "parent_id": "123",
                "is_dir": False,
                "size": 100,
                "hash_info": {"sha1": "same-hash"},
            },
            {
                "name": "会议录屏_2024.mp4",
                "file_id": "file-3",
                "parent_id": "123",
                "is_dir": False,
                "size": 200,
                "hash_info": {"sha1": "unique-hash"},
            },
        ]
    }
    return {"content": catalog.get(path, [])}


class FakeWriter:
    def __init__(self):
        self.scan_root_id = "123"
        self.items = {
            "123": {
                "file-1": "流浪地球2.2023.1080p.mkv",
                "file-2": "流浪地球2.2023.1080p.副本.mkv",
                "file-3": "会议录屏_2024.mp4",
            },
            "target-1": {},
        }
        self.created = []
        self.renamed = []
        self.moved = []

    @staticmethod
    def _item_id(item):
        return item["fid"]

    @staticmethod
    def _item_name(item):
        return item["fn"]

    def list_folder(self, folder_id):
        return [
            {"fid": file_id, "fn": name, "fc": "1"}
            for file_id, name in self.items.get(folder_id, {}).items()
        ]

    def ensure_directory(self, parts):
        self.created.append(tuple(parts))
        return "target-1"

    def find_child(self, parent_id, name):
        for file_id, current in self.items.get(parent_id, {}).items():
            if current.casefold() == name.casefold():
                return {"fid": file_id, "fn": current, "fc": "1"}
        return None

    def rename(self, file_id, new_name):
        for items in self.items.values():
            if file_id in items:
                items[file_id] = new_name
                self.renamed.append((file_id, new_name))
                return

    def move(self, file_id, destination_id):
        for items in self.items.values():
            if file_id in items:
                name = items.pop(file_id)
                self.items[destination_id][file_id] = name
                self.moved.append((file_id, destination_id))
                return

    def verify_item(self, parent_id, file_id, expected_name):
        return self.items.get(parent_id, {}).get(file_id) == expected_name


class ReportingAndOperationsTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        root = Path(self.tmpdir.name)
        env_path = root / ".env"
        env_path.write_text(
            "\n".join(
                [
                    "OPENLIST_BASE_URL=http://127.0.0.1:5244",
                    "OPENLIST_USERNAME=organizer-readonly",
                    "OPENLIST_PASSWORD=secret",
                    "ALLOWED_ROOT=/云下载",
                    "DEFAULT_SCAN_DIR=/云下载",
                    f"DB_PATH={root / 'index.sqlite'}",
                    f"LOG_PATH={root / 'organizer.log'}",
                    "WRITE_MODE=false",
                ]
            ),
            encoding="utf-8",
        )
        self.settings = load_settings(env_path)
        scan(
            self.settings,
            scan_dir="/云下载",
            max_depth=2,
            max_files=0,
            list_fn=fake_tree,
            scan_root_id="123",
        )

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_report_exports_html_excel_json_and_duplicate_risk(self):
        result = export_reports(self.settings, Path(self.tmpdir.name) / "reports")

        for key in ("html", "xlsx", "json"):
            self.assertTrue(Path(result[key]).is_file())
        payload = json.loads(Path(result["json"]).read_text(encoding="utf-8"))
        self.assertEqual(1, payload["duplicate_group_count"])
        self.assertEqual(2, payload["duplicate_file_count"])
        workbook = load_workbook(result["xlsx"], read_only=True)
        self.assertEqual(["总览", "整理计划"], workbook.sheetnames)
        workbook.close()
        self.assertIn("115 文件整理报告", Path(result["html"]).read_text(encoding="utf-8"))

    def test_safe_approval_excludes_duplicates(self):
        result = approve_safe_plans(self.settings)
        self.assertEqual(1, result["eligible"])
        conn = connect(self.settings.db_path)
        try:
            approved = conn.execute(
                "SELECT COUNT(*) AS n FROM organize_plans WHERE approved = 1"
            ).fetchone()["n"]
            self.assertEqual(1, approved)
        finally:
            conn.close()

    def test_manifest_blocks_duplicates_and_requires_confirmation(self):
        conn = connect(self.settings.db_path)
        try:
            ids = [row["id"] for row in conn.execute("SELECT id FROM organize_plans")]
            set_plan_approved(conn, ids, True)
            conn.commit()
        finally:
            conn.close()
        manifest = build_manifest(
            self.settings,
            scan_root_id="123",
            scan_root_path="/云下载",
        )
        self.assertEqual(1, manifest["operation_count"])
        self.assertEqual(2, manifest["blocked_count"])
        writer = FakeWriter()
        with self.assertRaises(OperationError):
            execute_manifest(self.settings, manifest, "WRONG", writer)
        self.assertEqual([], writer.moved)

        result = execute_manifest(
            self.settings,
            manifest,
            confirmation_code(manifest),
            writer,
        )
        self.assertEqual(1, result.succeeded)
        self.assertEqual(1, len(writer.moved))

    def test_tampered_manifest_is_rejected(self):
        approve_safe_plans(self.settings)
        manifest = build_manifest(
            self.settings,
            scan_root_id="123",
            scan_root_path="/云下载",
        )
        path = save_manifest(manifest, Path(self.tmpdir.name) / "manifest.json")
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload["organize_dir"] = "被修改"
        path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        with self.assertRaises(OperationError):
            load_manifest(path)

    def test_latest_successful_scan_hides_removed_remote_items(self):
        def reduced_tree(path: str):
            return {"content": [{
                "name": "会议录屏_2024.mp4",
                "file_id": "file-3",
                "parent_id": "123",
                "is_dir": False,
                "size": 200,
                "hash_info": {"sha1": "unique-hash"},
            }]}

        scan(
            self.settings,
            scan_dir="/云下载",
            max_depth=2,
            max_files=0,
            list_fn=reduced_tree,
            scan_root_id="123",
        )
        result = export_reports(self.settings, Path(self.tmpdir.name) / "reduced")
        payload = json.loads(Path(result["json"]).read_text(encoding="utf-8"))
        self.assertEqual(1, payload["stats"]["file_count"])
        self.assertEqual(1, len(payload["rows"]))
        self.assertEqual("会议录屏_2024.mp4", payload["rows"][0]["原文件名"])


if __name__ == "__main__":
    unittest.main()
