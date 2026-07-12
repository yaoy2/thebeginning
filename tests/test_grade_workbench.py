import pandas as pd
from openpyxl import load_workbook

from utils.grade_workbench import ScoreSettings, calculate_results, has_errors, round_score, validate_all
from utils.grade_workbench_export import build_review_workbook
from utils import grade_workbench_db


def sample_students():
    return pd.DataFrame([
        {
            "student_no": "S001",
            "name": "学生甲",
            "class_name": "测试班",
            "group_code": "第一组",
            "other_score": 80,
            "coefficient": 1.0,
            "individual_adjustment": 2,
            "adjustment_reason": "课堂贡献",
        },
        {
            "student_no": "S002",
            "name": "学生乙",
            "class_name": "测试班",
            "group_code": "第一组",
            "other_score": 80,
            "coefficient": 0.9,
            "individual_adjustment": 0,
            "adjustment_reason": "",
        },
    ])


def sample_groups():
    return pd.DataFrame([
        {
            "group_code": "第一组",
            "project_name": "虚拟项目",
            "pitch_score": 88,
            "report_score": 90,
            "group_adjustment": 1,
            "adjustment_reason": "统一展示奖励",
            "score_comment": "结构完整",
        }
    ])


def test_group_raw_scores_are_preserved_and_coefficient_is_applied():
    settings = ScoreSettings(global_adjustment=1)
    result = calculate_results(sample_students(), sample_groups(), settings)
    assert result.loc[0, "report_raw"] == 90
    assert result.loc[1, "report_raw"] == 90
    assert result.loc[0, "report_personal"] == 90
    assert result.loc[1, "report_personal"] == 81
    assert result.loc[0, "pitch_personal"] == 88
    assert result.loc[1, "pitch_personal"] == 79


def test_adjustments_do_not_change_group_scores():
    groups = sample_groups()
    original = groups[["pitch_score", "report_score"]].copy()
    calculate_results(sample_students(), groups, ScoreSettings(global_adjustment=5))
    pd.testing.assert_frame_equal(original, groups[["pitch_score", "report_score"]])


def test_half_up_rounding():
    assert round_score(80.5, "四舍五入为整数") == 81
    assert round_score(80.44, "保留一位小数") == 80.4


def test_validation_passes_for_complete_data():
    validation = validate_all(sample_students(), sample_groups(), ScoreSettings())
    assert not has_errors(validation)


def test_validation_blocks_missing_group_score():
    groups = sample_groups()
    groups.loc[0, "report_score"] = None
    validation = validate_all(sample_students(), groups, ScoreSettings())
    assert has_errors(validation)
    assert validation["message"].str.contains("报告原始分").any()


def test_global_adjustment_requires_a_reason_but_does_not_block_export():
    validation = validate_all(sample_students(), sample_groups(), ScoreSettings(global_adjustment=2))
    assert not has_errors(validation)
    assert validation["message"].str.contains("全体统一调整").any()


def test_review_workbook_preserves_raw_scores_and_adjustments(tmp_path):
    settings = ScoreSettings(global_adjustment=2, global_adjustment_reason="统一校准")
    students = sample_students()
    groups = sample_groups()
    results = calculate_results(students, groups, settings)
    validation = validate_all(students, groups, settings)
    destination = tmp_path / "review.xlsx"
    build_review_workbook(
        destination,
        {"name": "测试任务", "term": "测试学期", "course": "测试课程"},
        settings,
        students,
        groups,
        results,
        validation,
        pd.DataFrame([{"created_at": "2026-07-12", "action": "测试", "detail": "虚拟记录"}]),
    )
    workbook = load_workbook(destination, data_only=False)
    assert {"规则配置", "小组原始评分", "个人成绩计算", "校验摘要", "审计日志"}.issubset(workbook.sheetnames)
    group_sheet = workbook["小组原始评分"]
    assert group_sheet["C2"].value == 88
    assert group_sheet["D2"].value == 90
    result_sheet = workbook["个人成绩计算"]
    assert result_sheet["F2"].value == 88
    assert result_sheet["G2"].value == 90


def test_sqlite_task_persists_and_releases_database_file(tmp_path, monkeypatch):
    tasks_dir = tmp_path / "tasks"
    monkeypatch.setattr(grade_workbench_db, "TASKS_DIR", tasks_dir)
    task_id = grade_workbench_db.create_task("测试任务", "测试学期", "测试课程")
    grade_workbench_db.save_students(task_id, sample_students())
    grade_workbench_db.save_groups(task_id, sample_groups())
    loaded_students = grade_workbench_db.load_students(task_id)
    loaded_groups = grade_workbench_db.load_groups(task_id)
    assert len(loaded_students) == 2
    assert loaded_groups.loc[0, "report_score"] == 90
    database = grade_workbench_db.db_path(task_id)
    renamed = database.with_name("task-renamed.db")
    database.rename(renamed)
    assert renamed.exists()
