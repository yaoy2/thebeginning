import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils import report_grader


class ReportGraderTest(unittest.TestCase):
    def test_report_grader_page_uses_human_readable_table_headers(self):
        page_source = (Path(__file__).resolve().parents[1] / "pages" / "15_16_report_grader.py").read_text(
            encoding="utf-8"
        )

        for label in ["小组名", "报告标题", "起始行", "结束行", "报告字数", "识别成员数", "需复核权重数"]:
            self.assertIn(label, page_source)
        for label in ["学生姓名", "学号", "成员权重", "是否纳入", "个人报告基础分", "目标总成绩", "公式回算成绩"]:
            self.assertIn(label, page_source)
        self.assertIn('num_rows="dynamic"', page_source)

    def test_default_rubric_embeds_business_plan_score_table(self):
        rubric = report_grader.DEFAULT_BUSINESS_PLAN_RUBRIC

        self.assertIn("创业计划书评分标准（满分100分）", rubric)
        self.assertIn("项目可行性分析（15分）", rubric)
        self.assertIn("产品或服务的创新性（15分）", rubric)
        self.assertIn("市场机会及竞争（20分）", rubric)
        self.assertIn("营销策略（20分）", rubric)
        self.assertIn("财务分析的正确性及合理性（10分）", rubric)
        self.assertIn("团队管理（5分）", rubric)
        self.assertIn("风险分析与对策（5分）", rubric)
        self.assertIn("计划书内容完整、格式规范、AI使用规范（10分）", rubric)
        self.assertIn("合理使用AI工具辅助", rubric)
        self.assertEqual(100, sum(int(value) for value in __import__("re").findall(r"（(\d+)分）", rubric)))

    def test_splits_markdown_reports_and_extracts_members_with_weights(self):
        markdown = """
# 第一组 智慧养老项目

| 姓名 | 学号 | 权重 |
| --- | --- | --- |
| 张三 | 24032420501 | 1 |
| 李四 | 24032420502 | 0.9 |

正文内容 A

---

## 第二组 校园文创项目

王五 24032420503 权重 1.0
赵六 24032420504 权重 1.0

正文内容 B
"""

        reports = report_grader.split_markdown_reports(markdown)

        self.assertEqual([item.group_name for item in reports], ["第一组", "第二组"])
        self.assertEqual(reports[0].title, "第一组 智慧养老项目")
        self.assertGreater(reports[0].char_count, 20)
        self.assertEqual(
            [(m.name, m.student_id, m.weight, m.needs_review) for m in reports[0].members],
            [("张三", "24032420501", 1.0, False), ("李四", "24032420502", 0.9, False)],
        )

    def test_missing_member_weight_defaults_to_one_and_requires_review(self):
        members = report_grader.extract_members("张三 24032420501\n李四 24032420502 权重 0.8")

        self.assertEqual(members[0].weight, 1.0)
        self.assertTrue(members[0].needs_review)
        self.assertEqual(members[1].weight, 0.8)
        self.assertFalse(members[1].needs_review)

    def test_extracts_vertical_member_roster_blocks(self):
        text = """
商业计划书
成员名单

姓名
学号
专业、班级
权重系数

曾诗涵
24023420101
供应链管理，供应链24201
1

石静怡
24023420107
供应链管理，供应链24201
1

刘莹
24023420109
供应链管理，供应链24201
1

晏思雨
24023420112
供应链管理，供应链24201
1
"""

        members = report_grader.extract_members(text, "第5组")

        self.assertEqual(
            [(member.name, member.student_id, member.weight, member.needs_review) for member in members],
            [
                ("曾诗涵", "24023420101", 1.0, False),
                ("石静怡", "24023420107", 1.0, False),
                ("刘莹", "24023420109", 1.0, False),
                ("晏思雨", "24023420112", 1.0, False),
            ],
        )

    def test_vertical_roster_does_not_treat_year_labels_as_students(self):
        text = """
第四年
11115000
1
第五年
15600000
1
李欣雨
24052920113
电子商务，电商24201
0.95
"""

        members = report_grader.extract_members(text, "第二十三组")

        self.assertEqual([(member.name, member.student_id) for member in members], [("李欣雨", "24052920113")])

    def test_marks_student_id_length_outliers_for_review(self):
        text = """
张三 12345678 权重 1
李四 24052920113 权重 1
王五 24052920114 权重 1
"""

        members = report_grader.extract_members(text, "第一组")

        self.assertTrue(members[0].needs_review)
        self.assertFalse(members[1].needs_review)
        self.assertFalse(members[2].needs_review)

    def test_group_score_multiplies_member_weight_and_keeps_same_weight_equal(self):
        members = [
            report_grader.Member("第一组", "张三", "24032420501", 1.0, False),
            report_grader.Member("第一组", "李四", "24032420502", 0.9, False),
            report_grader.Member("第一组", "王五", "24032420503", 0.9, False),
        ]

        rows = report_grader.build_student_report_scores(
            {"第一组": report_grader.GroupScore("第一组", 90.0)}, members
        )

        self.assertEqual([row.report_score for row in rows], [90.0, 81.0, 81.0])
        self.assertEqual(report_grader.validate_same_weight_scores(rows), [])

    def test_component_linkage_keeps_report_score_when_target_final_is_overridden(self):
        row = report_grader.compose_grade_components(
            student_id="24032420501",
            name="张三",
            group_name="第一组",
            personal_report_score=80.0,
            bonus=0.0,
            target_final_score=87.0,
            reason="课堂贡献",
        )

        self.assertEqual(row.report_score, 80.0)
        self.assertEqual(row.target_final_score, 87.0)
        self.assertAlmostEqual(row.calculated_final_score, 87.0, places=1)
        self.assertAlmostEqual(report_grader.recalculate_final(row), 87.0, places=1)
        self.assertGreater(row.formative_average, row.terminal_average)
        self.assertEqual(row.addition, 0.0)

    def test_bonus_enters_final_score_when_no_manual_target_is_set(self):
        row = report_grader.compose_grade_components(
            student_id="24032420501",
            name="张三",
            group_name="第一组",
            personal_report_score=80.0,
            bonus=2.0,
            target_final_score=None,
            reason="竞赛加分",
        )

        self.assertEqual(row.target_final_score, 82.0)
        self.assertAlmostEqual(row.calculated_final_score, 82.0, places=1)
        self.assertEqual(row.addition, 2.0)

    def test_excel_writer_preserves_template_merge_and_writes_known_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "成绩单"
        sheet.merge_cells("A1:D1")
        headers = [
            "学号",
            "姓名",
            "组别",
            "课堂表现",
            "线上测试",
            "项目路演",
            "创业计划书",
            "加分项",
            "总成绩",
            "备注",
        ]
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=index, value=header)
        sheet.append(["24032420501", "张三", "1", None, None, None, None, None, None, None])
        source = BytesIO()
        workbook.save(source)

        row = report_grader.compose_grade_components(
            student_id="24032420501",
            name="张三",
            group_name="第一组",
            personal_report_score=80.0,
            bonus=2.0,
            target_final_score=87.0,
            reason="竞赛加分",
        )
        output = report_grader.fill_gradebook_template(source.getvalue(), [row])
        result = load_workbook(BytesIO(output))
        result_sheet = result["成绩单"]

        self.assertIn("A1:D1", [str(item) for item in result_sheet.merged_cells.ranges])
        self.assertEqual(result_sheet.cell(row=3, column=4).value, row.classroom)
        self.assertEqual(result_sheet.cell(row=3, column=8).value, 2.0)
        self.assertEqual(result_sheet.cell(row=3, column=9).value, 87.0)
        self.assertIn("竞赛加分", result_sheet.cell(row=3, column=10).value)


if __name__ == "__main__":
    unittest.main()
