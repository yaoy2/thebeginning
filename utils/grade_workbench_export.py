from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.grade_workbench import ScoreSettings


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")


def _write_frame(ws, frame: pd.DataFrame, start_row: int = 1) -> None:
    for col_idx, column in enumerate(frame.columns, 1):
        cell = ws.cell(start_row, col_idx, column)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(frame.itertuples(index=False, name=None), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx > 8)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(max(1, len(frame.columns)))}{max(start_row, start_row + len(frame))}"
    for col_idx, column in enumerate(frame.columns, 1):
        sample = [str(column)] + [str(value) for value in frame[column].head(80).fillna("")]
        width = min(32, max(10, max(len(value) for value in sample) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_review_workbook(
    destination: Path,
    meta: dict[str, str],
    settings: ScoreSettings,
    students: pd.DataFrame,
    groups: pd.DataFrame,
    results: pd.DataFrame,
    validation: pd.DataFrame,
    audit: pd.DataFrame,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "使用说明"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "教学评分审核工作簿"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    info = [
        ("任务名称", meta.get("name", "")),
        ("学期", meta.get("term", "")),
        ("课程", meta.get("course", "")),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("学生人数", len(students)),
        ("小组数量", len(groups)),
        ("校验状态", "存在错误，禁止作为正式成绩" if (not validation.empty and (validation["level"] == "错误").any()) else "硬校验通过"),
    ]
    for row, (label, value) in enumerate(info, 3):
        ws.cell(row, 1, label).fill = SECTION_FILL
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, value)
    ws["A12"] = "成绩计算口径"
    ws["A12"].fill = SECTION_FILL
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "最终成绩 = 其他考核成绩×其他权重 + 路演个人折算分×路演权重 + 报告个人折算分×报告权重 + 统一调整 + 小组调整 + 个人调整"
    ws.merge_cells("A13:F14")
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 34

    settings_frame = pd.DataFrame(
        [
            ["其他考核权重", settings.other_weight],
            ["路演权重", settings.pitch_weight],
            ["报告权重", settings.report_weight],
            ["统一调整", settings.global_adjustment],
            ["统一调整原因", settings.global_adjustment_reason],
            ["系数下限", settings.coefficient_min],
            ["系数上限", settings.coefficient_max],
            ["取整方式", settings.rounding],
            ["成绩下限", settings.score_floor],
            ["成绩上限", settings.score_cap],
        ],
        columns=["配置项", "当前值"],
    )
    _write_frame(wb.create_sheet("规则配置"), settings_frame)

    groups_export = groups.rename(columns={
        "group_code": "小组编号",
        "project_name": "项目名称",
        "pitch_score": "路演原始分",
        "report_score": "报告原始分",
        "group_adjustment": "小组调整",
        "adjustment_reason": "调整原因",
        "score_comment": "评分评语",
    })
    _write_frame(wb.create_sheet("小组原始评分"), groups_export)

    students_export = students.rename(columns={
        "student_no": "学号",
        "name": "姓名",
        "class_name": "班级",
        "group_code": "小组编号",
        "other_score": "其他考核成绩",
        "coefficient": "个人贡献系数",
        "individual_adjustment": "个人调整",
        "adjustment_reason": "个人调整原因",
    })
    _write_frame(wb.create_sheet("学生与个人调整"), students_export)

    result_names = {
        "student_no": "学号",
        "name": "姓名",
        "class_name": "班级",
        "group_code": "小组编号",
        "other_score": "其他考核成绩",
        "pitch_raw": "路演原始分",
        "report_raw": "报告原始分",
        "coefficient": "个人贡献系数",
        "pitch_personal": "路演个人折算分",
        "report_personal": "报告个人折算分",
        "global_adjustment": "统一调整",
        "group_adjustment": "小组调整",
        "individual_adjustment": "个人调整",
        "adjustment_total": "调整合计",
        "score_before_limit": "限幅前成绩",
        "final_score": "最终成绩",
        "adjustment_reason": "个人调整原因",
    }
    _write_frame(wb.create_sheet("个人成绩计算"), results.rename(columns=result_names))
    _write_frame(wb.create_sheet("校验摘要"), validation.rename(columns={"level": "级别", "category": "类别", "message": "说明"}))
    _write_frame(wb.create_sheet("审计日志"), audit.rename(columns={"created_at": "时间", "action": "操作", "detail": "说明"}))

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(destination)
    return destination
