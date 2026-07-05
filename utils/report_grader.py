from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import load_workbook


FORMATIVE_FIELDS = [
    "classroom",
    "online_test",
    "cost_budget",
    "org_risk",
    "legal_form",
    "mind_map",
    "business_canvas",
    "marketing_plan",
]

TERMINAL_FIELDS = ["roadshow", "business_plan"]

DEFAULT_BUSINESS_PLAN_RUBRIC = """创业计划书评分标准（满分100分）

1. 项目可行性分析（15分）
精确描述服务或产品的市场投资潜力、定位策略，分析项目实施后是否具备可行性，市场接受程度的调查资料是否完备。

2. 产品或服务的创新性（15分）
明确表述产品或服务的用户需求性、市场接受程度；指出产品或服务目前的技术及领先程度，以及市场需求的适应性。

3. 市场机会及竞争（20分）
明确阐述产品或服务的市场容量与趋势、市场竞争状况、市场变化趋势及潜力，细分目标市场及客户描述，估计市场份额和销售额，做到市场调查和分析的严密性。

4. 营销策略（20分）
准确描述市场进入策略和市场开发策略，包括竞争分析、市场细分、定位定价等，同时构建通畅合理的营销渠道和新颖、有吸引力的促销方式。

5. 财务分析的正确性及合理性（10分）
正确编制各类财务记账表，财务分析清晰明了，并能与计划配合实施。

6. 团队管理（5分）
合理进行营销、财务、行政、生产、技术团队分工，明确成员角色、组织结构、领导层分工、创业顾问及主要投资人，说明股份划分和战略实施能力。

7. 风险分析与对策（5分）
针对市场现状，客观分析项目面临的各种风险，并提出相应应对之策。

8. 计划书内容完整、格式规范、AI使用规范（10分）
创业计划所要求的各项内容无明显缺失，计划完整、全面、原创。语言通俗易懂，格式规范统一。合理使用AI工具辅助，不直接照搬AI生成内容。"""

GRADEBOOK_HEADER_ALIASES = {
    "student_id": ["学号"],
    "name": ["姓名"],
    "group": ["组别", "小组", "组号"],
    "classroom": ["课堂表现"],
    "online_test": ["线上测试", "在线测试"],
    "cost_budget": ["成本预算与财务计划", "成本预算"],
    "org_risk": ["组织架构与风险识别", "组织架构"],
    "legal_form": ["企业法律形态的选择", "法律形态"],
    "mind_map": ["思维导图"],
    "business_canvas": ["商业模式画布"],
    "marketing_plan": ["市场营销计划"],
    "roadshow": ["项目路演"],
    "business_plan": ["创业计划书"],
    "deduction": ["扣分项", "扣分"],
    "addition": ["加分项", "参与创新创业活动", "加分"],
    "final": ["总成绩", "课程总成绩"],
    "comment": ["备注"],
}


@dataclass(frozen=True)
class Member:
    group_name: str
    name: str
    student_id: str
    weight: float = 1.0
    needs_review: bool = False


@dataclass(frozen=True)
class ReportSection:
    title: str
    group_name: str
    content: str
    start_line: int
    end_line: int
    members: list[Member] = field(default_factory=list)

    @property
    def char_count(self) -> int:
        return len(self.content)


@dataclass(frozen=True)
class GroupScore:
    group_name: str
    report_score: float
    dimension_scores: dict[str, float] = field(default_factory=dict)
    comment: str = ""
    warnings: str = ""


@dataclass(frozen=True)
class StudentReportScore:
    group_name: str
    name: str
    student_id: str
    weight: float
    group_report_score: float
    report_score: float
    needs_review: bool = False


@dataclass(frozen=True)
class GradeRow:
    student_id: str
    name: str
    group_name: str
    report_score: float
    target_final_score: float
    calculated_final_score: float
    classroom: float
    online_test: float
    cost_budget: float
    org_risk: float
    legal_form: float
    mind_map: float
    business_canvas: float
    marketing_plan: float
    roadshow: float
    business_plan: float
    deduction: float = 0.0
    addition: float = 0.0
    reason: str = ""
    warnings: tuple[str, ...] = ()

    @property
    def formative_average(self) -> float:
        return round(sum(getattr(self, field_name) for field_name in FORMATIVE_FIELDS) / len(FORMATIVE_FIELDS), 2)

    @property
    def terminal_average(self) -> float:
        return round(sum(getattr(self, field_name) for field_name in TERMINAL_FIELDS) / len(TERMINAL_FIELDS), 2)


def split_markdown_reports(markdown: str) -> list[ReportSection]:
    lines = markdown.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    starts: list[int] = []
    for index, line in enumerate(lines):
        stripped = line.strip()
        if _is_report_heading(stripped):
            starts.append(index)

    if not starts:
        content = markdown.strip()
        if not content:
            return []
        title = _first_non_empty_line(lines) or "未命名报告"
        return [
            ReportSection(
                title=title,
                group_name=_extract_group_name(title) or "未识别小组",
                content=content,
                start_line=1,
                end_line=len(lines),
                members=extract_members(content, _extract_group_name(title) or "未识别小组"),
            )
        ]

    reports: list[ReportSection] = []
    for position, start in enumerate(starts):
        end = starts[position + 1] if position + 1 < len(starts) else len(lines)
        title = lines[start].lstrip("#").strip()
        content = "\n".join(lines[start:end]).strip()
        group_name = _extract_group_name(title) or f"报告{position + 1}"
        reports.append(
            ReportSection(
                title=title,
                group_name=group_name,
                content=content,
                start_line=start + 1,
                end_line=end,
                members=extract_members(content, group_name),
            )
        )
    return reports


def extract_members(text: str, group_name: str = "") -> list[Member]:
    table_members = _extract_members_from_tables(text, group_name)
    free_members = _extract_members_from_free_text(text, group_name)
    seen: set[str] = set()
    members: list[Member] = []
    for member in [*table_members, *free_members]:
        if member.student_id in seen:
            continue
        seen.add(member.student_id)
        members.append(member)
    return members


def parse_group_scores(raw_text: str) -> list[GroupScore]:
    text = raw_text.strip()
    if not text:
        return []
    if text.startswith("[") or text.startswith("{"):
        return _parse_group_scores_json(text)
    return _parse_group_scores_csv(text)


def build_student_report_scores(
    group_scores: dict[str, GroupScore],
    members: Iterable[Member],
) -> list[StudentReportScore]:
    rows: list[StudentReportScore] = []
    for member in members:
        score = group_scores.get(member.group_name)
        if score is None:
            rows.append(
                StudentReportScore(
                    group_name=member.group_name,
                    name=member.name,
                    student_id=member.student_id,
                    weight=member.weight,
                    group_report_score=0.0,
                    report_score=0.0,
                    needs_review=True,
                )
            )
            continue
        personal_score = round(score.report_score * member.weight, 1)
        rows.append(
            StudentReportScore(
                group_name=member.group_name,
                name=member.name,
                student_id=member.student_id,
                weight=member.weight,
                group_report_score=score.report_score,
                report_score=personal_score,
                needs_review=member.needs_review,
            )
        )
    return rows


def validate_same_weight_scores(rows: Iterable[StudentReportScore]) -> list[str]:
    buckets: dict[tuple[str, float], set[float]] = {}
    for row in rows:
        key = (row.group_name, round(row.weight, 3))
        buckets.setdefault(key, set()).add(round(row.report_score, 1))
    warnings: list[str] = []
    for (group_name, weight), scores in buckets.items():
        if len(scores) > 1:
            warnings.append(f"{group_name} 权重 {weight:g} 的成员报告分不一致：{sorted(scores)}")
    return warnings


def compose_grade_components(
    student_id: str,
    name: str,
    group_name: str,
    personal_report_score: float,
    bonus: float = 0.0,
    target_final_score: float | None = None,
    reason: str = "",
    deduction: float = 0.0,
) -> GradeRow:
    report_score = _round_score(personal_report_score)
    addition = _round_score(bonus)
    deduction_value = _round_score(deduction)
    target = _round_score(report_score + addition - deduction_value if target_final_score is None else target_final_score)
    target = min(100.0, max(0.0, target))

    terminal_average = report_score
    required_formative = (target + deduction_value - addition - terminal_average * 0.4) / 0.6
    warnings: list[str] = []
    if required_formative < 0 or required_formative > 100:
        warnings.append("目标总分超出当前报告分与加扣分可自然配平的范围，已把形成性分项限制在0-100。")
    formative_average = min(100.0, max(0.0, required_formative))

    formative_values = _spread_scores(formative_average, FORMATIVE_FIELDS)
    terminal_values = _spread_scores(terminal_average, TERMINAL_FIELDS, step=0.5)
    calculated = round(
        (sum(formative_values.values()) / len(FORMATIVE_FIELDS)) * 0.6
        + (sum(terminal_values.values()) / len(TERMINAL_FIELDS)) * 0.4
        - deduction_value
        + addition,
        1,
    )

    return GradeRow(
        student_id=str(student_id).strip(),
        name=str(name).strip(),
        group_name=str(group_name).strip(),
        report_score=report_score,
        target_final_score=target,
        calculated_final_score=calculated,
        deduction=deduction_value,
        addition=addition,
        reason=reason,
        warnings=tuple(warnings),
        **formative_values,
        **terminal_values,
    )


def recalculate_final(row: GradeRow) -> float:
    formative = sum(getattr(row, field_name) for field_name in FORMATIVE_FIELDS) / len(FORMATIVE_FIELDS)
    terminal = sum(getattr(row, field_name) for field_name in TERMINAL_FIELDS) / len(TERMINAL_FIELDS)
    return round(formative * 0.6 + terminal * 0.4 - row.deduction + row.addition, 1)


def fill_gradebook_template(template_bytes: bytes, rows: Iterable[GradeRow]) -> bytes:
    workbook = load_workbook(io.BytesIO(template_bytes))
    sheet = workbook.active
    header_map = _find_header_columns(sheet)
    row_map = _find_student_rows(sheet, header_map.get("student_id"))

    for grade_row in rows:
        target_row = row_map.get(_clean_student_id(grade_row.student_id))
        if not target_row:
            continue
        _write_known_columns(sheet, header_map, target_row, grade_row)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_ai_prompt(report: ReportSection, rubric: str) -> str:
    return f"""你现在是课程报告评审专家。请只根据报告文本内容，按同一把尺子评分。

评分标准：
{rubric.strip()}

请输出 CSV，表头固定为：
group_name,report_score,dimension_scores,comment,warnings

要求：
1. report_score 为小组报告基础分，0-100 分。
2. dimension_scores 用 JSON 对象表达各维度分数。
3. 不要给同组成员分别打不同报告分；成员个人分由系统按权重换算。
4. comment 简短说明主要依据，warnings 标注材料缺失或需人工复核点。

报告标题：{report.title}
识别小组：{report.group_name}

报告全文：
{report.content}
""".strip()


def rows_to_csv(rows: Iterable[dict]) -> str:
    rows = list(rows)
    if not rows:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def _is_report_heading(line: str) -> bool:
    if not line.startswith("#"):
        return False
    title = line.lstrip("#").strip()
    if len(title) < 2:
        return False
    return bool(re.search(r"(第\s*[一二三四五六七八九十\d]+\s*组|小组|项目|报告)", title))


def _extract_group_name(title: str) -> str:
    match = re.search(r"第\s*([一二三四五六七八九十\d]+)\s*组", title)
    if match:
        return f"第{match.group(1)}组"
    match = re.search(r"([\w\u4e00-\u9fff]{1,20}小组)", title)
    if match:
        return match.group(1)
    return ""


def _first_non_empty_line(lines: list[str]) -> str:
    for line in lines:
        stripped = line.strip().lstrip("#").strip()
        if stripped:
            return stripped
    return ""


def _extract_members_from_tables(text: str, group_name: str) -> list[Member]:
    members: list[Member] = []
    lines = text.splitlines()
    for index, line in enumerate(lines):
        if "|" not in line:
            continue
        cells = _split_markdown_row(line)
        if not {"姓名", "学号"}.issubset(set(cells)):
            continue
        header = cells
        weight_index = _header_index(header, ["权重", "weight"])
        name_index = _header_index(header, ["姓名"])
        id_index = _header_index(header, ["学号"])
        for row_line in lines[index + 1 :]:
            if "|" not in row_line:
                break
            row = _split_markdown_row(row_line)
            if not row or all(re.fullmatch(r"-+", cell) for cell in row):
                continue
            if len(row) <= max(name_index, id_index):
                continue
            student_id = _clean_student_id(row[id_index])
            if not student_id:
                continue
            weight_text = row[weight_index] if weight_index is not None and len(row) > weight_index else ""
            weight, needs_review = _parse_weight(weight_text)
            members.append(Member(group_name, row[name_index].strip(), student_id, weight, needs_review))
    return members


def _extract_members_from_free_text(text: str, group_name: str) -> list[Member]:
    members: list[Member] = []
    pattern = re.compile(
        r"(?P<name>[\u4e00-\u9fff]{2,4})[^\n\d]{0,12}(?P<id>\d{8,14})(?P<tail>[^\n]{0,30})"
    )
    for match in pattern.finditer(text):
        tail = match.group("tail")
        weight_match = re.search(r"(?:权重|weight)\s*[:：]?\s*(\d+(?:\.\d+)?)", tail, re.I)
        if weight_match:
            weight, needs_review = _parse_weight(weight_match.group(1))
        else:
            weight, needs_review = 1.0, True
        members.append(Member(group_name, match.group("name"), _clean_student_id(match.group("id")), weight, needs_review))
    return members


def _split_markdown_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def _header_index(header: list[str], names: list[str]) -> int | None:
    lowered = [item.lower() for item in header]
    for name in names:
        for index, item in enumerate(lowered):
            if name.lower() in item:
                return index
    return None


def _parse_weight(value: object) -> tuple[float, bool]:
    text = str(value or "").strip()
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return 1.0, True
    return float(match.group(0)), False


def _parse_group_scores_csv(text: str) -> list[GroupScore]:
    cleaned = re.sub(r"^```(?:csv)?|```$", "", text.strip(), flags=re.I | re.M).strip()
    reader = csv.DictReader(io.StringIO(cleaned))
    scores: list[GroupScore] = []
    for row in reader:
        group_name = (row.get("group_name") or row.get("小组") or row.get("组名") or "").strip()
        if not group_name:
            continue
        score = float(row.get("report_score") or row.get("报告分") or 0)
        dimension_scores = _load_dimension_scores(row.get("dimension_scores") or "")
        scores.append(
            GroupScore(
                group_name=group_name,
                report_score=score,
                dimension_scores=dimension_scores,
                comment=row.get("comment", ""),
                warnings=row.get("warnings", ""),
            )
        )
    return scores


def _parse_group_scores_json(text: str) -> list[GroupScore]:
    payload = json.loads(text)
    items = payload if isinstance(payload, list) else [payload]
    return [
        GroupScore(
            group_name=str(item.get("group_name", "")).strip(),
            report_score=float(item.get("report_score", 0)),
            dimension_scores=dict(item.get("dimension_scores") or {}),
            comment=str(item.get("comment", "")),
            warnings=str(item.get("warnings", "")),
        )
        for item in items
        if item.get("group_name")
    ]


def _load_dimension_scores(value: str) -> dict[str, float]:
    try:
        payload = json.loads(value) if value else {}
    except json.JSONDecodeError:
        return {}
    return {str(key): float(score) for key, score in payload.items()}


def _spread_scores(average: float, fields: list[str], step: float = 0.1) -> dict[str, float]:
    values = {field: _round_score(average) for field in fields}
    if not fields:
        return values
    current_total = sum(values.values())
    target_total = round(average * len(fields), 1)
    difference = round(target_total - current_total, 1)
    index = 0
    while abs(difference) >= 0.1 and index < len(fields) * 20:
        field = fields[index % len(fields)]
        direction = step if difference > 0 else -step
        next_value = values[field] + direction
        if 0 <= next_value <= 100:
            values[field] = round(next_value, 1)
            difference = round(target_total - sum(values.values()), 1)
        index += 1
    return values


def _round_score(value: float) -> float:
    return round(float(value), 1)


def _find_header_columns(sheet) -> dict[str, int]:
    found: dict[str, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value or "").strip()
            if not value:
                continue
            for key, aliases in GRADEBOOK_HEADER_ALIASES.items():
                if key in found:
                    continue
                if any(alias in value for alias in aliases):
                    found[key] = cell.column
    return found


def _find_student_rows(sheet, student_id_column: int | None) -> dict[str, int]:
    if not student_id_column:
        return {}
    rows: dict[str, int] = {}
    for row_index in range(1, sheet.max_row + 1):
        student_id = _clean_student_id(sheet.cell(row=row_index, column=student_id_column).value)
        if student_id:
            rows[student_id] = row_index
    return rows


def _write_known_columns(sheet, header_map: dict[str, int], row_index: int, grade_row: GradeRow) -> None:
    values = {
        "group": grade_row.group_name,
        "classroom": grade_row.classroom,
        "online_test": grade_row.online_test,
        "cost_budget": grade_row.cost_budget,
        "org_risk": grade_row.org_risk,
        "legal_form": grade_row.legal_form,
        "mind_map": grade_row.mind_map,
        "business_canvas": grade_row.business_canvas,
        "marketing_plan": grade_row.marketing_plan,
        "roadshow": grade_row.roadshow,
        "business_plan": grade_row.business_plan,
        "deduction": grade_row.deduction,
        "addition": grade_row.addition,
        "final": grade_row.target_final_score,
        "comment": grade_row.reason,
    }
    for key, value in values.items():
        column = header_map.get(key)
        if column:
            sheet.cell(row=row_index, column=column, value=value)


def _clean_student_id(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"\d{8,14}", text)
    return match.group(0) if match else ""
